#!/usr/bin/env python3
"""L18 - Load Turkish fiscal data from SBB, MoF, World Bank, and TurkStat HDARP.

Parses:
  1. SBB consolidated budget (1975-2007) — % share breakdown
  2. MoF General Budget Balance (2006-2025) — annual TL values
  3. World Bank structural data (1980-2019) — GDP, growth, structure
  4. TurkStat HDARP extraction — compensation of employees & labor share (1980-2006)

Outputs: data/raw-data/parsed/turkey_*.csv
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

import pandas as pd
import numpy as np
from utils.paths import PARSED_RAW, INPUTS, ensure_dirs

TURKEY_DIR = INPUTS / "ExternalSources" / "Turkey2022"
MUHASEBAT_DIR = TURKEY_DIR / "muhasebat"


def _load_sbb_consolidated():
    """Parse SBB consolidated budget Excel (3 sections: 1975-85, 86-95, 97-2007)."""
    path = TURKEY_DIR / "sbb_konsolide_butce_gelir_harcama.xlsx"
    if not path.exists():
        return pd.DataFrame()

    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    sections = [
        (6, 9, 34),   # Row 6 has years, data rows 9-34
        (41, 44, 69),  # Section 2
        (75, 78, 103), # Section 3
    ]

    all_data = []
    for yr_row, data_start, data_end in sections:
        year_cells = {}
        for c in range(3, ws.max_column + 1):
            v = ws.cell(row=yr_row, column=c).value
            if v:
                yr_str = str(v).strip().split()[0].replace("(1)", "").strip()
                try:
                    year_cells[int(yr_str)] = c
                except ValueError:
                    pass

        for year, col in year_cells.items():
            row_data = {"year": year}
            for r in range(data_start, min(data_end, ws.max_row + 1)):
                label = ws.cell(row=r, column=2).value
                val = ws.cell(row=r, column=col).value
                if label and val is not None:
                    label = str(label).strip()
                    try:
                        row_data[label] = float(val)
                    except (ValueError, TypeError):
                        pass
            if len(row_data) > 1:
                all_data.append(row_data)

    if not all_data:
        return pd.DataFrame()

    df = pd.DataFrame(all_data).set_index("year").sort_index()
    df = df[~df.index.duplicated(keep="last")]
    return df


def _load_mof_budget_balance():
    """Parse MoF General Budget Balance 2006-2025 (20 yearly sheets, monthly→annual)."""
    path = MUHASEBAT_DIR / "General-Budget-Balance-and-Financing-2006-2025.xls"
    if not path.exists():
        return pd.DataFrame()

    try:
        import xlrd
    except ImportError:
        return pd.DataFrame()

    wb = xlrd.open_workbook(path)
    all_years = []

    for sheet_name in wb.sheet_names():
        try:
            year = int(sheet_name)
        except ValueError:
            continue

        sh = wb.sheet_by_name(sheet_name)
        annual = {"year": year}

        for r in range(sh.nrows):
            label = str(sh.cell_value(r, 0)).strip() if sh.ncols > 0 else ""
            if not label:
                continue

            # Get the last column (TOTAL or December cumulative)
            total_col = sh.ncols - 1
            for c in range(sh.ncols - 1, 0, -1):
                v = sh.cell_value(r, c)
                if v and str(v).strip() not in ("", "TOTAL", "Total"):
                    try:
                        float(v)
                        total_col = c
                        break
                    except (ValueError, TypeError):
                        pass

            val = sh.cell_value(r, total_col)
            try:
                val = float(val)
            except (ValueError, TypeError):
                continue

            if "GENERAL BUDGET REVENUES" in label.upper():
                annual["total_revenue_ml_tl"] = val
            elif label.upper().startswith("TAX REVENUES") or label == "Tax Revenues":
                annual["tax_revenue_ml_tl"] = val
            elif "Direct Taxes" in label:
                annual["direct_taxes_ml_tl"] = val
            elif "Indirect Taxes" in label:
                annual["indirect_taxes_ml_tl"] = val
            elif "GENERAL BUDGET EXPENDITURES" in label.upper():
                annual["total_expenditure_ml_tl"] = val

        if len(annual) > 1:
            all_years.append(annual)

    if not all_years:
        return pd.DataFrame()

    return pd.DataFrame(all_years).set_index("year").sort_index()


def load():
    ensure_dirs()
    steps = []
    outputs = []

    # SBB consolidated budget
    sbb = _load_sbb_consolidated()
    if not sbb.empty:
        out = PARSED_RAW / "turkey_sbb_consolidated.csv"
        sbb.to_csv(out)
        outputs.append(str(out))
        steps.append(f"SBB consolidated: {len(sbb)} years ({sbb.index.min()}-{sbb.index.max()}), {len(sbb.columns)} columns")
        print(f"    [L18] SBB consolidated: {len(sbb)} years")

    # MoF budget balance
    mof = _load_mof_budget_balance()
    if not mof.empty:
        out = PARSED_RAW / "turkey_mof_budget_balance.csv"
        mof.to_csv(out)
        outputs.append(str(out))
        steps.append(f"MoF budget balance: {len(mof)} years ({mof.index.min()}-{mof.index.max()})")
        print(f"    [L18] MoF budget: {len(mof)} years")

    # World Bank structural data
    wb_path = TURKEY_DIR / "worldbank_turkey_structural_1980_2019.csv"
    if wb_path.exists():
        wb_df = pd.read_csv(wb_path, index_col=0)
        out = PARSED_RAW / "turkey_worldbank_structural.csv"
        wb_df.to_csv(out)
        outputs.append(str(out))
        steps.append(f"World Bank structural: {len(wb_df)} years")
        print(f"    [L18] World Bank: {len(wb_df)} years")

    # World Bank fiscal
    wbf_path = TURKEY_DIR / "worldbank_turkey_fiscal_1980_2019.csv"
    if wbf_path.exists():
        wbf = pd.read_csv(wbf_path, index_col=0)
        out = PARSED_RAW / "turkey_worldbank_fiscal.csv"
        wbf.to_csv(out)
        outputs.append(str(out))
        steps.append(f"World Bank fiscal: {len(wbf)} years")
        print(f"    [L18] World Bank fiscal: {len(wbf)} years")

    # TurkStat HDARP: compensation of employees & labor share (Table 20.35/20.37)
    ts_path = PARSED_RAW / "turkstat_compensation_labor_share_1980_2006.csv"
    if ts_path.exists():
        ts_df = pd.read_csv(ts_path, skiprows=2, index_col=0)
        steps.append(f"TurkStat HDARP compensation: {len(ts_df)} years ({ts_df.index.min()}-{ts_df.index.max()})")
        print(f"    [L18] TurkStat HDARP: {len(ts_df)} years (compensation + labor share)")

    # FRED: Turkey labor share (Penn World Table via FRED, 1950-2019)
    try:
        from dotenv import load_dotenv
        env_path = Path(__file__).resolve().parent.parent.parent / "data" / "user-inputs" / "api_keys.env"
        if env_path.exists():
            load_dotenv(env_path)
        from utils.fetchers.fred_fetcher import fetch_fred
        fred_data = fetch_fred("LABSHPTRA156NRUG", start_date="1980-01-01", end_date="2019-12-31")
        obs = fred_data.get("observations", [])
        if obs:
            fred_rows = []
            for o in obs:
                try:
                    yr = int(o["date"][:4])
                    val = float(o["value"]) / 100.0  # convert from percent to ratio
                    fred_rows.append({"year": yr, "labor_share_fred": val})
                except (ValueError, KeyError):
                    continue
            if fred_rows:
                fred_df = pd.DataFrame(fred_rows).set_index("year")
                fred_out = PARSED_RAW / "turkey_fred_labor_share_1980_2019.csv"
                fred_df.to_csv(fred_out)
                outputs.append(str(fred_out))
                steps.append(f"FRED Turkey labor share: {len(fred_df)} years ({fred_df.index.min()}-{fred_df.index.max()})")
                print(f"    [L18] FRED Turkey: {len(fred_df)} years (LABSHPTRA156NRUG)")
    except Exception as e:
        steps.append(f"FRED Turkey: failed ({e})")

    return {
        "series_id": "L18",
        "status": "ok" if outputs else "fail",
        "steps": steps,
        "outputs": outputs,
    }
