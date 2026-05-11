#!/usr/bin/env python3
"""
Anu Chopped Converter v1.0
==========================

Converts Shaikh Chopped Excel files into the Anu Chopped self-documenting
CSV format:
    Row 1: Per-column metadata (source, methodology, units, coverage)
    Row 2: Subseries IDs (S001A, S001B, ..., S001)
    Row 3+: Data with Year in column 1

Usage:
    python convert_to_chopped.py <input_dir> <output_dir> [--catalog catalog.json]
    python convert_to_chopped.py single <input_file> <output_file>

Created: 2026-02-11
Standard: Anu Chopped v1.0
"""

import sys
import csv
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# CONFIGURATION
# =============================================================================

APPENDIX_TO_CHAPTER = {
    'Appendix2': 2, 'Appendix5': 5, 'Appendix6': 6, 'Appendix7': 7,
    'Appendix8': 8, 'Appendix9': 9, 'Appendix10': 10, 'Appendix11': 11,
    'Appendix12': 12, 'Appendix14': 14, 'Appendix15': 15, 'Appendix16': 16,
    'Appendix17': 17,
}

WIDE_FORMAT_FILES = {
    'Appendix6_Table68I1.xlsx', 'Appendix6_Table68I2.xlsx',
    'Appendix6_Table68I3.xlsx', 'Appendix6_Table68II1.xlsx',
    'Appendix6_Table68II2.xlsx', 'Appendix6_Table68II3.xlsx',
    'Appendix6_Table68II4.xlsx', 'Appendix6_Table68II5.xlsx',
    'Appendix6_Table68II6.xlsx', 'Appendix6_Table68II7.xlsx',
    'Appendix2_GDPperCapita.xlsx',
}

MATRIX_FORMAT_FILES = {
    'Appendix9_1947fixed.xlsx', 'Appendix9_1958fixed.xlsx',
    'Appendix9_1963fixed.xlsx', 'Appendix9_1967fixed.xlsx',
    'Appendix9_1972fixed.xlsx', 'Appendix9_1998Fixed.xlsx',
    'Appendix9_1998Circ.xlsx', 'Appendix9_ReswitchExamples.xlsx',
    'Appendix9_ReswitchingPseudoProductionFunction.xlsx',
    'Appendix9_pvdevexample.xlsx',
    'Appendix8_Bain42IndustryProfit.xlsx',
    'Appendix8_Bain42IndustryAggregates.xlsx',
    'Appendix8_CorrectedBainData.xlsx',
    'Appendix8_DemsetzRatesOfReturn.xlsx',
    'Appendix8_Semmler19843.3.xlsx',
    'Appendix8_StiglerRatesOfProfit.xlsx',
    'Appendix7_SalterTable9.xlsx', 'Appendix7_SalterTable10.xlsx',
    'Appendix7_SalterULCPriceTable28.xlsx',
    'Appendix7_SalterULCPriceTable33.xlsx',
    'Appendix15_WorldInflationDataByCountry.xlsx',
    'Appendix15_WorldInflationDataLambda.xlsx',
    'Appendix17_USIRS2011.xlsx',
}

DOCUMENTATION_FILES = {
    'Appendix5_Documentation.xlsx', 'Appendix6_Contents.xlsx',
    'Appendix7_Documentation.xlsx', 'Appendix9_Documentation.xlsx',
    'Appendix10_Documentation.xlsx', 'Appendix11_Documentation.xlsx',
    'Appendix14_Documentation.xlsx', 'Appendix15_Documentation.xlsx',
}

# Subseries ID counter -- tracks allocation per series
_subseries_counters: Dict[str, int] = {}


def get_chapter(filename: str) -> Optional[int]:
    """Get chapter number from filename."""
    for prefix, ch in APPENDIX_TO_CHAPTER.items():
        if filename.startswith(prefix):
            return ch
    return None


def get_chapter_dir(chapter: int) -> str:
    """Get chapter directory name."""
    return f"ch{chapter:02d}"


def classify_source(text: str) -> str:
    """Classify source text into a source type."""
    if not isinstance(text, str) or len(text) < 3:
        return 'unknown'
    t = text.lower()
    if any(x in t for x in ['nipa', 'national income', 't 1.', 't 2.', 'table 1.']):
        return 'bea_nipa'
    if any(x in t for x in ['fixed assets', 'fa ', 'fa6', 'table 6.']):
        return 'bea_fixed_assets'
    if any(x in t for x in ['frb', 'federal reserve', 'g-17', 'h.15']):
        return 'frb'
    if any(x in t for x in ['bls', 'bureau of labor']):
        return 'bls'
    if 'fred' in t:
        return 'fred'
    if any(x in t for x in ['measuring worth', 'measuringworth']):
        return 'measuring_worth'
    if any(x in t for x in ['ibbotson', 'sbbi']):
        return 'ibbotson'
    if 'shiller' in t:
        return 'shiller'
    if any(x in t for x in ['nber', 'national bureau']):
        return 'nber'
    if any(x in t for x in ['census', 'historical statistics']):
        return 'census'
    if 'oecd' in t:
        return 'oecd'
    if any(x in t for x in ['imf', 'ifs']):
        return 'imf'
    if 'world bank' in t:
        return 'world_bank'
    if any(x in t for x in ['penn world', 'pwt']):
        return 'penn_world'
    if any(x in t for x in ['economic report', 'erp', 'president']):
        return 'econ_report_pres'
    if any(x in t for x in ['irs', 'internal revenue']):
        return 'irs'
    authors = ['ayres', 'jastram', 'macaulay', 'kendrick', 'salter', 'bain',
               'stigler', 'semmler', 'demsetz', 'harberger', 'maddison']
    if any(a in t for a in authors):
        return 'historical_book'
    if 'long term eco' in t or 'bea' in t:
        return 'historical_book'
    if any(x in t for x in ['calculated', 'derived', 'computed', 'sum of']):
        return 'calculated'
    return 'unknown'


def extract_base_year(text: str) -> Tuple[Optional[int], Optional[int]]:
    """Extract base year and value from metadata text."""
    if not isinstance(text, str):
        return None, None
    patterns = [
        r'indexed to (\d{4})\s*=\s*(\d+)',
        r'(\d{4})\s*=\s*(\d+)',
        r'reindexed to (\d{4})',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            groups = match.groups()
            base_year = int(groups[0])
            base_value = int(groups[1]) if len(groups) > 1 and groups[1] else 100
            if 1800 < base_year < 2020:
                return base_year, base_value
    return None, None


def safe_str(val) -> str:
    """Convert value to string, handling None/NaN."""
    if val is None:
        return ''
    s = str(val)
    if s in ('nan', 'None', 'NaT', 'NaN'):
        return ''
    return s.strip()


def safe_float(val) -> Optional[float]:
    """Convert value to float, returning None for non-numeric."""
    if val is None:
        return None
    try:
        f = float(val)
        if f != f:  # NaN check
            return None
        return f
    except (ValueError, TypeError):
        return None


def is_year(val) -> bool:
    """Check if a value looks like a year."""
    f = safe_float(val)
    return f is not None and 1600 < f < 2100


def parse_quarter(val) -> Optional[str]:
    """Parse quarterly date like '1980Q1' or '1980q1'. Returns the string or None."""
    s = safe_str(val)
    if re.match(r'^\d{4}[Qq][1-4]$', s):
        return s.upper()
    return None


# =============================================================================
# PATTERN A: STANDARD TIME SERIES
# =============================================================================

def convert_long_format(filepath: Path) -> Dict[str, Any]:
    """
    Convert a standard time-series Excel file to Anu Chopped format.
    Returns dict with 'rows' (list of row lists), 'catalog' (metadata dict).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheet = wb.active

    all_rows = []
    for row in sheet.iter_rows(values_only=True):
        all_rows.append(list(row))
    wb.close()

    if len(all_rows) < 3:
        return {'rows': [], 'catalog': {}, 'error': 'Too few rows'}

    # Row 0 = metadata string(s)
    row0 = all_rows[0]
    row0_text = ' '.join(safe_str(v) for v in row0 if safe_str(v))

    # Row 1 = headers
    headers = [safe_str(v) for v in all_rows[1]]

    # Find the year column index and detect quarterly data
    year_idx = None
    is_quarterly = False
    for i, h in enumerate(headers):
        if h.lower() in ('year', 'yr', 'date', 'time period'):
            year_idx = i
            break
    if year_idx is None:
        # Try first column
        first_vals = [safe_float(all_rows[r][0]) for r in range(2, min(10, len(all_rows)))]
        if any(is_year(v) for v in first_vals if v is not None):
            year_idx = 0

    # Check for quarterly data (e.g., "1980Q1")
    if year_idx is not None:
        sample_vals = [safe_str(all_rows[r][year_idx]) for r in range(2, min(10, len(all_rows))) if r < len(all_rows)]
        if any(parse_quarter(v) for v in sample_vals):
            is_quarterly = True

    if year_idx is None:
        return {'rows': [], 'catalog': {}, 'error': 'No year column found'}

    # Data columns (everything except year)
    data_col_indices = [i for i in range(len(headers)) if i != year_idx and headers[i]]

    if not data_col_indices:
        return {'rows': [], 'catalog': {}, 'error': 'No data columns found'}

    # Build per-column metadata from row 0
    per_col_meta = []
    for ci in data_col_indices:
        meta_parts = []
        # Get the cell from row 0 above this column
        if ci < len(row0) and safe_str(row0[ci]):
            meta_parts.append(safe_str(row0[ci]))
        if not meta_parts:
            # Fall back to the overall metadata
            meta_parts.append(row0_text[:200])
        per_col_meta.append(' '.join(meta_parts))

    # Allocate subseries IDs
    stem = filepath.stem
    chapter = get_chapter(filepath.name)
    subseries_ids = allocate_subseries_ids(stem, len(data_col_indices), headers, data_col_indices)

    # Build output rows
    # Row 1: metadata
    date_label = 'Quarter' if is_quarterly else 'Year'
    meta_row = [date_label] + per_col_meta
    # Row 2: IDs
    id_row = [''] + subseries_ids
    # Row 3+: data
    data_rows = []
    year_min, year_max = 9999, 0
    for r in range(2, len(all_rows)):
        raw = all_rows[r]
        raw_val = raw[year_idx] if year_idx < len(raw) else None

        if is_quarterly:
            q = parse_quarter(raw_val)
            if q is None:
                continue
            # Extract year part for range tracking
            yr = int(q[:4])
            year_min = min(year_min, yr)
            year_max = max(year_max, yr)
            row_data = [q]
        else:
            year_val = safe_float(raw_val)
            if year_val is None or not is_year(year_val):
                continue
            year = int(year_val)
            year_min = min(year_min, year)
            year_max = max(year_max, year)
            row_data = [year]

        for ci in data_col_indices:
            val = safe_float(raw[ci] if ci < len(raw) else None)
            row_data.append(val if val is not None else '')
        data_rows.append(row_data)

    # Sort by index
    data_rows.sort(key=lambda r: str(r[0]))

    output_rows = [meta_row, id_row] + data_rows

    # Build catalog entry
    columns_catalog = {}
    for idx, (sid, ci) in enumerate(zip(subseries_ids, data_col_indices)):
        col_meta = per_col_meta[idx] if idx < len(per_col_meta) else ''
        base_yr, base_val = extract_base_year(col_meta)
        col_name = headers[ci] if ci < len(headers) else f'col_{ci}'
        is_final = not re.match(r'^S\d{3}[A-Z]', sid)
        columns_catalog[sid] = {
            'name': col_name,
            'description': col_meta[:300],
            'source_type': classify_source(col_meta),
            'type': 'final' if is_final else 'raw',
            'coverage': [year_min, year_max],
            'base_year': base_yr,
            'base_value': base_val,
        }

    catalog = {
        'source_excel': filepath.name,
        'chapter': chapter,
        'format': 'time_series',
        'year_range': [year_min, year_max] if data_rows else None,
        'row_count': len(data_rows),
        'columns': columns_catalog,
        'linked_figures': [],
        'linked_series': [],
    }

    return {'rows': output_rows, 'catalog': catalog}


def allocate_subseries_ids(stem: str, n_cols: int, headers: List[str],
                           col_indices: List[int]) -> List[str]:
    """Allocate subseries IDs for data columns."""
    # For now, use a simple sequential allocation: S###A, S###B, ..., S###
    # In production, these would be looked up from DEFINITIVE_SERIES_CATALOG
    base_id = _get_series_base(stem)
    ids = []
    for i in range(n_cols):
        if i == n_cols - 1:
            # Last column = final series (no letter suffix)
            ids.append(base_id)
        else:
            ids.append(f"{base_id}{_idx_to_letter(i)}")
    return ids


def _idx_to_letter(i: int) -> str:
    """Convert index to letter suffix: 0->A, 1->B, ..., 25->Z, 26->AA, 27->AB, ..."""
    if i < 26:
        return chr(ord('A') + i)
    else:
        # For > 26 columns, use AA, AB, AC, ...
        first = chr(ord('A') + (i // 26) - 1)
        second = chr(ord('A') + (i % 26))
        return f"{first}{second}"


def _get_series_base(stem: str) -> str:
    """Map a file stem to a series base ID."""
    # Known mappings from the Capitalism Data project
    KNOWN = {
        'Appendix2_IndustrialProduction': 'S001',
        'Appendix2_RealInvestmentUS_1832-2010': 'S002',
        'Appendix2_MeasuringWorthGDP_1889-2010': 'S003',
        'Appendix2_Ayres': 'S004',
        'Appendix2_ManufacturingProductivityAndRealWages1889-2010': 'S007',
        'Appendix2_ManufacturingProductivity': 'S008',
        'Appendix2_Unemployment': 'S009',
        'Appendix2_GDPperCapita': 'S017',
        'Appendix5_DATALRprices': 'S010',
        'Appendix6_Table68II7': 'S013',
        'Appendix10_Ibbotson': 'S040',
        'Appendix10_IntroPPrice': 'S041',
        'Appendix10_USLR': 'S042',
        'Appendix12_CreditInflUnempl': 'S065',
        'Appendix15_MeasuringWorthCPI': 'S072',
        'Appendix15_USInflation': 'S073',
        'Appendix15_Argentina': 'S074',
        'Appendix16_ProfitRates': 'S090',
        'Appendix16_WageProdData': 'S091',
        'Appendix16_DebtIncRatio': 'S092',
        'Appendix16_HouseholdDebtService': 'S093',
        'Appendix16_RXRRULCOECD': 'S094',
        'Appendix17_USIRS2011': 'S100',
    }
    if stem in KNOWN:
        return KNOWN[stem]

    # Generate a hash-based ID for unknown files
    if stem not in _subseries_counters:
        # Allocate from S200+ range to avoid collision
        _subseries_counters[stem] = 200 + len(_subseries_counters)
    return f"S{_subseries_counters[stem]:03d}"


# =============================================================================
# PATTERN B: WIDE FORMAT (years as columns)
# =============================================================================

def convert_wide_format(filepath: Path) -> Dict[str, Any]:
    """
    Convert a wide-format Excel file (years as columns, variables as rows)
    to Anu Chopped format by transposing.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheet = wb.active

    all_rows = []
    for row in sheet.iter_rows(values_only=True):
        all_rows.append(list(row))
    wb.close()

    if len(all_rows) < 3:
        return {'rows': [], 'catalog': {}, 'error': 'Too few rows'}

    # Row 0 = table title/metadata
    row0_text = safe_str(all_rows[0][0]) if all_rows[0] else ''

    # Row 1 = headers (Table, Description, Source, Variable, 1947, 1948, ...)
    headers = [safe_str(v) for v in all_rows[1]]

    # Find year columns
    year_cols = []  # (col_index, year_value)
    for i, h in enumerate(headers):
        f = safe_float(h)
        if f is not None and 1700 < f < 2100:
            year_cols.append((i, int(f)))

    if len(year_cols) < 3:
        return {'rows': [], 'catalog': {}, 'error': 'Not enough year columns'}

    # Label columns are everything before the first year column
    label_end = year_cols[0][0]
    label_headers = headers[:label_end]

    # Extract years
    years = [yc[1] for yc in year_cols]
    year_min, year_max = min(years), max(years)

    # Each data row (row 2+) becomes a column in the output
    col_metas = []
    col_ids = []
    col_data = []  # list of lists, one per variable

    base_id = _get_series_base(filepath.stem)

    for row_idx in range(2, len(all_rows)):
        raw = all_rows[row_idx]
        if not raw or all(v is None for v in raw):
            continue

        # Extract labels
        labels = {}
        for li, lh in enumerate(label_headers):
            if li < len(raw):
                labels[lh] = safe_str(raw[li])

        # Determine variable name
        var_name = labels.get('Variable', '') or labels.get('Description', '')
        if not var_name:
            for lh in label_headers:
                if labels.get(lh):
                    var_name = labels[lh]
                    break
        if not var_name:
            continue

        # Build metadata
        source = labels.get('Source', '')
        description = labels.get('Description', '')
        table_ref = labels.get('Table', '')
        meta = f"{description}; Source: {source}" if source else description
        if table_ref:
            meta = f"[{table_ref}] {meta}"
        meta = meta.strip('; ') or var_name

        # Extract data for this variable
        values = []
        for ci, yr in year_cols:
            val = safe_float(raw[ci] if ci < len(raw) else None)
            values.append(val if val is not None else '')

        # Only include if there's meaningful data
        non_empty = sum(1 for v in values if v != '')
        if non_empty < 2:
            continue

        col_metas.append(meta[:300])
        col_data.append(values)

        # Allocate ID
        var_idx = len(col_ids)
        col_ids.append(f"{base_id}{_idx_to_letter(var_idx)}")

    if not col_data:
        return {'rows': [], 'catalog': {}, 'error': 'No data variables found'}

    # Build output rows
    meta_row = ['Year'] + col_metas
    id_row = [''] + col_ids
    data_rows = []
    for yi, yr in enumerate(years):
        row = [yr]
        for cd in col_data:
            row.append(cd[yi] if yi < len(cd) else '')
        data_rows.append(row)

    output_rows = [meta_row, id_row] + data_rows

    # Build catalog
    columns_catalog = {}
    for sid, meta in zip(col_ids, col_metas):
        columns_catalog[sid] = {
            'name': sid,
            'description': meta,
            'source_type': classify_source(meta),
            'type': 'raw',
            'coverage': [year_min, year_max],
            'base_year': None,
            'base_value': None,
        }

    catalog = {
        'source_excel': filepath.name,
        'chapter': get_chapter(filepath.name),
        'format': 'wide_table',
        'year_range': [year_min, year_max],
        'row_count': len(data_rows),
        'columns': columns_catalog,
        'linked_figures': [],
        'linked_series': [],
    }

    return {'rows': output_rows, 'catalog': catalog}


# =============================================================================
# PATTERN C: MATRIX / CROSS-SECTIONAL
# =============================================================================

def convert_matrix_format(filepath: Path) -> Dict[str, Any]:
    """
    Convert a cross-sectional/matrix Excel file to Anu Chopped format.
    Keeps original row structure but adds metadata and ID rows.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheet = wb.active

    all_rows = []
    for row in sheet.iter_rows(values_only=True):
        all_rows.append(list(row))
    wb.close()

    if len(all_rows) < 3:
        return {'rows': [], 'catalog': {}, 'error': 'Too few rows'}

    # Row 0 = metadata
    row0_text = safe_str(all_rows[0][0]) if all_rows[0] else ''

    # Row 1 = headers
    headers = [safe_str(v) for v in all_rows[1]]

    # Index column is the first column
    index_label = headers[0] if headers else 'Index'

    # Data columns
    data_headers = headers[1:]
    base_id = _get_series_base(filepath.stem)

    # Build per-column metadata from row 0
    per_col_meta = []
    col_ids = []
    for i, h in enumerate(data_headers):
        meta = f"{row0_text[:100]}; Column: {h}" if row0_text else h
        per_col_meta.append(meta[:300])
        col_ids.append(f"FPR{get_chapter(filepath.name) or 0:03d}_C{i + 1}")

    # Build output rows
    meta_row = [index_label] + per_col_meta
    id_row = [''] + col_ids

    data_rows = []
    for r in range(2, len(all_rows)):
        raw = all_rows[r]
        if not raw or all(v is None for v in raw):
            continue
        row_out = [safe_str(raw[0]) if raw else '']
        for ci in range(1, len(headers)):
            val = safe_float(raw[ci] if ci < len(raw) else None)
            row_out.append(val if val is not None else '')
        data_rows.append(row_out)

    output_rows = [meta_row, id_row] + data_rows

    # Catalog
    columns_catalog = {}
    for sid, meta, h in zip(col_ids, per_col_meta, data_headers):
        columns_catalog[sid] = {
            'name': h,
            'description': meta,
            'source_type': classify_source(row0_text),
            'type': 'cross_sectional',
            'coverage': None,
            'base_year': None,
            'base_value': None,
        }

    catalog = {
        'source_excel': filepath.name,
        'chapter': get_chapter(filepath.name),
        'format': 'cross_sectional',
        'year_range': None,
        'row_count': len(data_rows),
        'columns': columns_catalog,
        'linked_figures': [],
        'linked_series': [],
    }

    return {'rows': output_rows, 'catalog': catalog}


# =============================================================================
# WRITER
# =============================================================================

def write_chopped_csv(rows: List[List], filepath: Path):
    """Write rows to an Anu Chopped CSV file."""
    filepath.parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        for row in rows:
            writer.writerow(row)


# =============================================================================
# MAIN CONVERSION ORCHESTRATOR
# =============================================================================

def detect_format(filepath: Path) -> str:
    """Detect file format."""
    name = filepath.name
    if name in DOCUMENTATION_FILES:
        return 'documentation'
    if name in MATRIX_FORMAT_FILES:
        return 'matrix'
    if name in WIDE_FORMAT_FILES:
        return 'wide'
    return 'long'


def convert_file(input_path: Path, output_dir: Path) -> Optional[Dict]:
    """
    Convert a single Excel file to Anu Chopped CSV.
    Returns catalog entry or None on skip/error.
    """
    fmt = detect_format(input_path)
    chapter = get_chapter(input_path.name)

    if fmt == 'documentation':
        print(f"  SKIP (documentation): {input_path.name}")
        return {
            'source_excel': input_path.name,
            'chapter': chapter,
            'format': 'documentation',
            'year_range': None,
            'row_count': 0,
            'columns': {},
            'linked_figures': [],
            'linked_series': [],
            'note': 'Documentation-only file, no data converted',
        }

    print(f"  Converting ({fmt}): {input_path.name}")

    try:
        if fmt == 'long':
            result = convert_long_format(input_path)
        elif fmt == 'wide':
            result = convert_wide_format(input_path)
        elif fmt == 'matrix':
            result = convert_matrix_format(input_path)
        else:
            result = convert_long_format(input_path)
    except Exception as e:
        print(f"  ERROR: {input_path.name}: {e}")
        return None

    if result.get('error'):
        print(f"  WARN: {input_path.name}: {result['error']}")
        return None

    if not result.get('rows'):
        print(f"  WARN: {input_path.name}: No output rows")
        return None

    # Determine output path
    ch_dir = get_chapter_dir(chapter) if chapter else 'other'
    out_name = input_path.stem + '.csv'
    out_path = output_dir / ch_dir / out_name

    write_chopped_csv(result['rows'], out_path)
    print(f"    -> {out_path.relative_to(output_dir)} ({result['catalog'].get('row_count', 0)} rows)")

    return result['catalog']


def convert_all(input_dir: Path, output_dir: Path, catalog_path: Optional[Path] = None):
    """Convert all Excel files in input_dir to Anu Chopped CSVs."""
    xlsx_files = sorted(input_dir.glob('Appendix*.xlsx'))
    # Skip temp files
    xlsx_files = [f for f in xlsx_files if not f.name.startswith('~$')]

    print(f"Anu Chopped Converter v1.0")
    print(f"Input:  {input_dir}")
    print(f"Output: {output_dir}")
    print(f"Files:  {len(xlsx_files)}")
    print("=" * 60)

    catalog = {
        'version': '1.0',
        'standard': 'Anu Chopped v1.0',
        'project': 'CD2',
        'generated': datetime.now().strftime('%Y-%m-%d'),
        'source_location': str(output_dir),
        'total_files': 0,
        'total_columns': 0,
        'files': {},
    }

    success = 0
    skipped = 0
    errors = 0

    for xlsx in xlsx_files:
        entry = convert_file(xlsx, output_dir)
        if entry:
            ch_dir = get_chapter_dir(entry.get('chapter') or 0)
            key = f"{ch_dir}/{xlsx.stem}.csv"
            catalog['files'][key] = entry
            if entry.get('format') != 'documentation':
                success += 1
                catalog['total_columns'] += len(entry.get('columns', {}))
            else:
                skipped += 1
        else:
            errors += 1

    catalog['total_files'] = success

    print("=" * 60)
    print(f"Done: {success} converted, {skipped} skipped, {errors} errors")
    print(f"Total columns: {catalog['total_columns']}")

    # Write catalog
    if catalog_path is None:
        catalog_path = output_dir.parent / 'ANU_CHOPPED_CATALOG.json'
    catalog_path.parent.mkdir(parents=True, exist_ok=True)
    with open(catalog_path, 'w', encoding='utf-8') as f:
        json.dump(catalog, f, indent=2, ensure_ascii=False)
    print(f"Catalog: {catalog_path}")

    return catalog


# =============================================================================
# CLI
# =============================================================================

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage:")
        print("  python convert_to_chopped.py <input_dir> <output_dir>")
        print("  python convert_to_chopped.py single <input.xlsx> <output.csv>")
        sys.exit(1)

    if sys.argv[1] == 'single':
        inp = Path(sys.argv[2])
        out = Path(sys.argv[3])
        result = convert_file(inp, out.parent)
        if result:
            print(json.dumps(result, indent=2))
    else:
        inp = Path(sys.argv[1])
        out = Path(sys.argv[2])
        cat = Path(sys.argv[3]) if len(sys.argv) > 3 else None
        convert_all(inp, out, cat)
