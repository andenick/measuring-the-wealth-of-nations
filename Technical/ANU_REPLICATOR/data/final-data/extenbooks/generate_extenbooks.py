"""Generate Extenbook XLSX files for all 26 Wave 1 series."""

import json
import csv
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Paths
BASE = Path(__file__).resolve().parents[3] / "Technical"
REGISTRY_PATH = BASE / "series_registry.json"
RESEARCH_DIR = BASE / "research"
SERIES_DIR = BASE / "ANU_REPLICATOR/data/final-data/series"
OUTPUT_DIR = BASE / "ANU_REPLICATOR/data/final-data/extenbooks"

# Wave 1 series IDs
WAVE1_IDS = (
    [f"T{n}" for n in range(501, 517)] +
    [f"T{n}" for n in range(601, 610)] +
    ["T901"]
)

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
LABEL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def style_header_row(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)


def create_data_sheet(wb, series_id):
    """Sheet 1: Data - copy from source CSV."""
    ws = wb.active
    ws.title = "Data"

    csv_path = SERIES_DIR / f"{series_id}.csv"
    if not csv_path.exists():
        ws.append(["year", "book", "combined"])
        ws.append(["[No CSV found]", "", ""])
        style_header_row(ws, 3)
        auto_width(ws)
        return

    with open(csv_path, "r", newline="") as f:
        reader = csv.reader(f)
        headers = next(reader)
        ws.append(headers)
        for row in reader:
            typed = []
            for val in row:
                if val == "":
                    typed.append(None)
                else:
                    try:
                        typed.append(float(val))
                    except ValueError:
                        typed.append(val)
            ws.append(typed)

    style_header_row(ws, len(headers))
    auto_width(ws)


def create_provenance_sheet(wb, series_id, info):
    """Sheet 2: Provenance metadata."""
    ws = wb.create_sheet("Provenance")

    ws.append(["Field", "Value"])
    style_header_row(ws, 2)

    year_range = info.get("year_range", [])
    yr_str = f"{year_range[0]}-{year_range[1]}" if len(year_range) == 2 else str(year_range)

    rows = [
        ("series_id", series_id),
        ("series_name", info.get("name", "")),
        ("chapter", info.get("chapter", "")),
        ("book_table", info.get("book_table", "")),
        ("year_range", yr_str),
        ("wave", 1),
        ("status", info.get("status", "")),
        ("generated", "2026-03-21"),
    ]
    for field, value in rows:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=field).font = LABEL_FONT
        ws.cell(row=r, column=2, value=str(value))

    auto_width(ws)


def create_research_sheet(wb, series_id):
    """Sheet 3: Research entries from JSON."""
    ws = wb.create_sheet("Research")
    ws.append(["entry_type", "content", "source_ref"])
    style_header_row(ws, 3)

    json_path = RESEARCH_DIR / f"{series_id}_research.json"
    if not json_path.exists():
        ws.append(["[No research JSON found]", "", ""])
        auto_width(ws)
        return

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    for entry in data.get("entries", []):
        ws.append([
            entry.get("entry_type", ""),
            entry.get("content", ""),
            entry.get("source_ref", ""),
        ])

    # Wrap text for content column
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    auto_width(ws)
    # Override content column to fixed width for readability
    ws.column_dimensions["B"].width = 80


def create_construction_sheet(wb, series_id, info):
    """Sheet 4: Construction steps."""
    ws = wb.create_sheet("Construction")
    ws.append(["step", "operation", "input", "output", "notes"])
    style_header_row(ws, 5)

    steps = info.get("construction", [])
    if not steps:
        ws.append(["[No construction steps defined]", "", "", "", ""])
        auto_width(ws)
        return

    for s in steps:
        step_num = s.get("step", "")
        op = s.get("op", "")

        # Gather input info
        inputs = []
        if "subseries" in s:
            inputs.extend(s["subseries"])
        if "inputs" in s:
            inputs.extend(s["inputs"] if isinstance(s["inputs"], list) else [s["inputs"]])
        input_str = ", ".join(str(i) for i in inputs)

        output = s.get("output", "")

        # Build notes from remaining keys
        notes_parts = []
        if "formula" in s:
            notes_parts.append(f"formula: {s['formula']}")
        if "at_year" in s:
            notes_parts.append(f"splice_at: {s['at_year']}")
        if "method" in s:
            notes_parts.append(f"method: {s['method']}")
        if "source" in s:
            notes_parts.append(f"source: {s['source']}")
        notes = "; ".join(notes_parts)

        ws.append([step_num, op, input_str, output, notes])

    auto_width(ws)


def generate_extenbook(series_id, info):
    """Generate a single extenbook XLSX."""
    wb = Workbook()
    create_data_sheet(wb, series_id)
    create_provenance_sheet(wb, series_id, info)
    create_research_sheet(wb, series_id)
    create_construction_sheet(wb, series_id, info)

    out_path = OUTPUT_DIR / f"{series_id}_extenbook.xlsx"
    wb.save(str(out_path))
    return out_path


def main():
    with open(REGISTRY_PATH, "r", encoding="utf-8") as f:
        registry = json.load(f)

    series_map = registry.get("series", {})
    created = []
    errors = []

    for sid in WAVE1_IDS:
        if sid not in series_map:
            errors.append(f"{sid}: not found in registry")
            continue
        try:
            path = generate_extenbook(sid, series_map[sid])
            created.append(sid)
            print(f"  OK  {sid} -> {path.name}")
        except Exception as e:
            errors.append(f"{sid}: {e}")
            print(f"  ERR {sid}: {e}")

    print(f"\nGenerated {len(created)}/{len(WAVE1_IDS)} extenbooks.")
    if errors:
        print("Errors:")
        for e in errors:
            print(f"  {e}")


if __name__ == "__main__":
    main()
