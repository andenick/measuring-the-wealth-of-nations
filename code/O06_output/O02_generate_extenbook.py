"""O02 — Generate Anu Extenbook xlsx for each implemented series.

Per Anu Extenbook Standard v3.2, each workbook has 4 sheets:
  Sheet 1: Data         — year-indexed columns (subseries + final)
  Sheet 2: Provenance   — DPR text + cross-refs
  Sheet 3: Research     — research JSON entries (one row per entry)
  Sheet 4: Construction — registry construction steps + identity checks

This generator reads `series_registry.json` + `docs/series/{sid}_DPR.md` +
`research/{sid}_research.json` + `data/final/{sid}.csv` and writes
`extenbooks/{sid}.xlsx`. Uses openpyxl with the framework's color standard.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from utils.paths import DATA_FINAL, DOCS_SERIES, EXTENBOOKS_DIR, REGISTRY, RESEARCH_DIR  # noqa: E402


# Color standard from anu-extenbook spec
FILL_HEADER    = PatternFill("solid", fgColor="4472C4")
FILL_METADATA  = PatternFill("solid", fgColor="D9E2F3")
FILL_ORIGINAL  = PatternFill("solid", fgColor="FFF2CC")
FILL_FINAL     = PatternFill("solid", fgColor="E6FFE6")
FILL_NAN       = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT    = Font(name="Calibri", bold=True, color="FFFFFF")
META_FONT      = Font(name="Calibri", italic=True, size=9)


def _autosize(ws, min_w: int = 10, max_w: int = 60) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_w
        for cell in col:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, min(len(str(v)), max_w))
        ws.column_dimensions[col_letter].width = max_len + 2


def write_data_sheet(ws, sid: str, entry: dict) -> None:
    ws.title = "Data"
    final_csv = DATA_FINAL / f"{sid}.csv"
    if not final_csv.exists():
        ws.append([f"No final/{sid}.csv yet — series status: {entry.get('status','?')}"])
        return

    df = pd.read_csv(final_csv)
    pivot = (df.pivot_table(index="year", columns="series_id", values="value", aggfunc="first")
               .reset_index().sort_values("year"))

    name = entry.get("name", "")
    units = entry.get("units", "")
    subseries = entry.get("subseries", {})

    # Row 1: metadata
    ws.append([f"{sid} — {name}"] + [
        f"{subseries.get(c,{}).get('source','')}, units={units}"
        for c in pivot.columns if c != "year"
    ])
    for cell in ws[1]:
        cell.fill = FILL_METADATA
        cell.font = META_FONT

    # Row 2: column headers
    headers = ["Year"] + [c for c in pivot.columns if c != "year"]
    ws.append(headers)
    for cell in ws[2]:
        cell.fill = FILL_HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Row 3+: data
    for _, row in pivot.iterrows():
        ws.append([int(row["year"])] + [row[c] for c in pivot.columns if c != "year"])

    # Color the subseries columns (book period yellow, final green)
    final_col = None
    for i, h in enumerate(headers, start=1):
        if h == sid:                          # bare series ID = final
            final_col = i
    n_rows = ws.max_row
    for r in range(3, n_rows + 1):
        for c in range(2, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None or cell.value == "":
                cell.fill = FILL_NAN
            elif c == final_col:
                cell.fill = FILL_FINAL
            else:
                cell.fill = FILL_ORIGINAL

    _autosize(ws)


def write_provenance_sheet(ws, sid: str, entry: dict, dpr_path: Path) -> None:
    ws.title = "Provenance"
    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = HEADER_FONT

    rows: list[tuple[str, str]] = [
        ("Series ID",     sid),
        ("Name",          entry.get("name", "")),
        ("Chapter",       str(entry.get("chapter", ""))),
        ("Book Table",    entry.get("book_table", "")),
        ("Units",         entry.get("units", "")),
        ("Content Type",  entry.get("content_type", "")),
        ("Year Range",    f"{entry.get('year_range', [None,None])[0]}-{entry.get('year_range', [None,None])[1]}"),
        ("Status",        entry.get("status", "")),
        ("Figures",       ", ".join(entry.get("figures", []))),
        ("Loader",        entry.get("loader", "") or ""),
        ("Processor",     entry.get("processor", "") or ""),
        ("Wave",          str(entry.get("wave", ""))),
    ]
    for k, v in rows:
        ws.append([k, v])

    # Proxy disclosure section (Decision 0001 + the project proxy-disclosure rule)
    if entry.get("proxy") is True:
        ws.append([])
        ws.append(["PROXY DISCLOSURE", ""])
        for cell in ws[ws.max_row]:
            cell.fill = FILL_HEADER
            cell.font = HEADER_FONT
        ws.append(["proxy", "TRUE — this series is NOT a direct replication of the book concept."])
        ws.append(["proxy_basis", str(entry.get("proxy_basis") or "<not specified>")])
        ws.append(["proxy_disclosure", str(entry.get("proxy_disclosure") or "Matrix-derived or surrogate construction; consult docs/series/" + sid + "_DPR.md for full disclosure.")])
        if entry.get("proxy_real_fix_plan"):
            ws.append(["proxy_real_fix_plan", str(entry["proxy_real_fix_plan"])])

    # KB sources / provenance index
    if entry.get("kb_sources"):
        ws.append([])
        ws.append(["KB Sources", ""])
        for src in entry.get("kb_sources", []):
            ws.append(["", str(src)])

    # Subseries provenance
    if entry.get("subseries"):
        ws.append([])
        ws.append(["Subseries", "source / role"])
        for cell in ws[ws.max_row]:
            cell.fill = FILL_HEADER
            cell.font = HEADER_FONT
        for sub_id, sub_meta in entry["subseries"].items():
            src = sub_meta.get("source", "")
            role = sub_meta.get("role", sub_meta.get("description", ""))
            ws.append([sub_id, f"{src} | {role}" if role else src])

    # Append DPR text as a single multi-line block
    if dpr_path.exists():
        ws.append([])
        ws.append(["DPR (markdown)", ""])
        dpr_text = dpr_path.read_text(encoding="utf-8")
        for line in dpr_text.splitlines():
            ws.append(["", line])
    _autosize(ws)


def write_research_sheet(ws, sid: str, research_path: Path) -> None:
    ws.title = "Research"
    if not research_path.exists():
        ws.append([f"No research JSON for {sid}"])
        return

    r = json.loads(research_path.read_text(encoding="utf-8"))
    ws.append(["entry_id_or_type", "content_or_quote", "source_ref"])
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = HEADER_FONT

    # Verbatim quotes first (Decision 0001 emphasizes these)
    verbatim_first: list = []
    other_entries: list = []
    for e in r.get("entries", []):
        et = e.get("entry_type", e.get("type", ""))
        if et == "verbatim_quote":
            verbatim_first.append(e)
        else:
            other_entries.append(e)
    for e in verbatim_first + other_entries:
        eid = e.get("entry_id", e.get("entry_type", e.get("type", "")))
        content = e.get("content", e.get("quote", ""))
        ref = e.get("source_ref", e.get("source_location", e.get("kb_reference", "")))
        ws.append([str(eid), str(content), str(ref)])

    # Top-level verbatim_quotes block (alternative format)
    if r.get("verbatim_quotes"):
        ws.append([])
        ws.append(["VERBATIM QUOTES (top-level)", "", ""])
        for cell in ws[ws.max_row]:
            cell.fill = FILL_HEADER
            cell.font = HEADER_FONT
        for q in r["verbatim_quotes"]:
            if isinstance(q, dict):
                ws.append([
                    q.get("id", q.get("quote_id", "")),
                    q.get("quote", q.get("content", "")),
                    q.get("source_ref", q.get("source_location", q.get("page", ""))),
                ])
            else:
                ws.append(["", str(q), ""])

    ws.append([])
    ws.append(["methodology_summary", r.get("methodology_summary", "")])
    if r.get("known_issues"):
        ws.append([])
        ws.append(["known_issues:"])
        for issue in r["known_issues"]:
            ws.append(["", issue])
    if r.get("cross_references"):
        ws.append([])
        ws.append(["cross_references:"])
        for x in r["cross_references"]:
            ws.append(["", x])
    if r.get("citations"):
        ws.append([])
        ws.append(["citations:"])
        for c in r["citations"]:
            ws.append([c.get("id", c.get("citation_id", "")), c.get("full", c.get("full_citation", ""))])
    _autosize(ws)


def write_construction_sheet(ws, sid: str, entry: dict) -> None:
    ws.title = "Construction"
    ws.append(["step", "op", "inputs/subseries", "output", "formula/at_year"])
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = HEADER_FONT
    for s in entry.get("construction", []):
        ws.append([
            s.get("step", ""),
            s.get("op", ""),
            ", ".join(s.get("inputs", []) or s.get("subseries", []) or []),
            s.get("output", ""),
            s.get("formula", "") or s.get("at_year", "") or s.get("method", ""),
        ])

    # Proxy disclosure in Construction sheet (mirrors Provenance for accountability)
    if entry.get("proxy") is True:
        ws.append([])
        ws.append(["PROXY CONSTRUCTION NOTICE", ""])
        for cell in ws[ws.max_row]:
            cell.fill = FILL_HEADER
            cell.font = HEADER_FONT
        ws.append(["proxy", "TRUE"])
        ws.append(["proxy_basis", str(entry.get("proxy_basis") or "<not specified>")])
        ws.append(["proxy_disclosure", str(entry.get("proxy_disclosure") or "Series is a proxy, not a direct replication. See DPR.")])
        if entry.get("proxy_real_fix_plan"):
            ws.append(["real_fix_plan", str(entry["proxy_real_fix_plan"])])

    if entry.get("extension"):
        ws.append([])
        ws.append(["EXTENSION", ""])
        for cell in ws[ws.max_row]:
            cell.fill = FILL_HEADER
            cell.font = HEADER_FONT
        for k, v in entry["extension"].items():
            ws.append([k, json.dumps(v) if isinstance(v, (dict, list)) else str(v)])
    elif entry.get("extension") is None:
        ws.append([])
        ws.append(["EXTENSION", "null - series does not extend beyond book period"])

    val = entry.get("validation", {})
    if val:
        ws.append([])
        ws.append(["VALIDATION", ""])
        for k, v in val.items():
            ws.append([k, json.dumps(v) if isinstance(v, (dict, list)) else str(v)])
    _autosize(ws)


def build_workbook(sid: str, entry: dict) -> Path | None:
    final_csv = DATA_FINAL / f"{sid}.csv"
    if not final_csv.exists():
        return None
    wb = Workbook()
    write_data_sheet(wb.active, sid, entry)
    write_provenance_sheet(wb.create_sheet("Provenance"), sid, entry, DOCS_SERIES / f"{sid}_DPR.md")
    write_research_sheet(wb.create_sheet("Research"), sid, RESEARCH_DIR / f"{sid}_research.json")
    write_construction_sheet(wb.create_sheet("Construction"), sid, entry)

    EXTENBOOKS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = EXTENBOOKS_DIR / f"{sid}.xlsx"
    wb.save(out_path)
    return out_path


def run() -> dict:
    reg = json.loads(REGISTRY.read_text(encoding="utf-8"))
    written = []
    skipped = []
    for sid, entry in reg["series"].items():
        out = build_workbook(sid, entry)
        if out is None:
            skipped.append(sid)
        else:
            written.append(sid)
    print(f"    [O02] Extenbooks: wrote {len(written)} workbooks; skipped {len(skipped)} (no final data yet)")
    return {"written": written, "skipped": skipped}


if __name__ == "__main__":
    run()
