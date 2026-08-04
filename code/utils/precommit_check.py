"""Pre-commit check for RMWND Anu Framework Decisions 0001-0008.

Standalone, dependency-light enforcement of the eight binding decisions that
govern the RMWND replication project. Designed to be runnable either as a git
pre-commit hook (`--staged`) or manually across the whole project (`--all`).

Usage
-----
    # Whole-project audit (default if no git repo)
    python Technical/code/utils/precommit_check.py --all

    # Staged-file mode (git pre-commit hook default)
    python Technical/code/utils/precommit_check.py --staged

    # Ad-hoc subset (glob patterns relative to project root)
    python Technical/code/utils/precommit_check.py --files "Technical/chopped/*.csv"

Exit code: 0 on all PASS / WARN, 1 on any FAIL.

Authored 2026-05-23 for v1.1 Phase 2 framework infrastructure. Read-only with
respect to series_registry.json, PIPELINE_STATE.json, LEDGER, MANIFEST, and
SUBSERIES_PLAN.json — this script never writes data artifacts.

The eight checks
----------------
- Decision 0001: extenbook 4-sheet (or 5-sheet w/ Validation) canonical layout.
- Decision 0002: every registry series has non-empty validation.reference_values
  (allowed-empty only when derived_statistics is non-empty, per 0008).
- Decision 0003: extension binary invariant
  * extension == null  -> status MUST NOT be 'validated_book_and_extension' (FAIL)
  * extension != null  -> status SHOULD contain 'extension' (WARN — project may
    legitimately defer the upgrade until extension is actually executed; see
    the project conventions).
- Decision 0004: L01 / P02 / V03 compact naming `<phase>_<SID>[_<suffix>].py`.
- Decision 0005: chopped CSV wide format (row 1 # metadata, row 2 'Year', row 3+ data).
- Decision 0006: code is source of truth (heuristic always-PASS; documentation).
- Decision 0007: research JSON has >= 3 canonical verbatim_quote entries with
  required fields (entry_type, content/verbatim_quote, source_ref).
- Decision 0008: validation.reference_values keys parse as int years; values
  are finite floats. derived_statistics is unconstrained.
"""
from __future__ import annotations

import argparse
import fnmatch
import json
import math
import re
import subprocess
import sys
from pathlib import Path
from typing import Iterable

# ---------------------------------------------------------------------------
# Layout
# ---------------------------------------------------------------------------
# This file: Technical/code/utils/precommit_check.py
SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[3]  # bundle/project root
TECHNICAL_ROOT = PROJECT_ROOT / "Technical"
REGISTRY_PATH = TECHNICAL_ROOT / "series_registry.json"

CANONICAL_EXTENBOOK_SHEETS = ["Data", "Provenance", "Research", "Construction"]
ALLOWED_EXTENBOOK_SHEETS_5 = CANONICAL_EXTENBOOK_SHEETS + ["Validation"]

# Phase prefixes for Decision 0004
PHASE_PREFIXES = ("L01_", "P02_", "V03_", "M04_", "A05_", "O06_", "S00_")

# ---------------------------------------------------------------------------
# Result accumulator
# ---------------------------------------------------------------------------
class Result:
    __slots__ = ("decision", "level", "path", "message")
    def __init__(self, decision: str, level: str, path: str, message: str) -> None:
        self.decision = decision
        self.level = level  # PASS | WARN | FAIL
        self.path = path
        self.message = message

    def fmt(self) -> str:
        return f"[{self.level:4}] {self.decision:13} {self.path:60} {self.message}"


class Report:
    def __init__(self) -> None:
        self.results: list[Result] = []

    def add(self, decision: str, level: str, path: str, message: str) -> None:
        self.results.append(Result(decision, level, path, message))

    def counts(self) -> dict[str, int]:
        c = {"PASS": 0, "WARN": 0, "FAIL": 0}
        for r in self.results:
            c[r.level] = c.get(r.level, 0) + 1
        return c

    def exit_code(self) -> int:
        return 1 if any(r.level == "FAIL" for r in self.results) else 0


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------
def _git_staged_files() -> list[Path]:
    try:
        cp = subprocess.run(
            ["git", "diff", "--cached", "--name-only", "--diff-filter=ACM"],
            cwd=str(PROJECT_ROOT),
            capture_output=True,
            text=True,
            check=False,
        )
        if cp.returncode != 0:
            return []
        return [PROJECT_ROOT / p.strip() for p in cp.stdout.splitlines() if p.strip()]
    except FileNotFoundError:
        return []


def _has_git_repo() -> bool:
    return (PROJECT_ROOT / ".git").exists() or (PROJECT_ROOT.parent / ".git").exists()


def _glob_files(pattern: str) -> list[Path]:
    return [p for p in PROJECT_ROOT.glob(pattern) if p.is_file()]


def discover_files(args: argparse.Namespace) -> tuple[list[Path], str]:
    """Return (files, mode_label)."""
    if args.files:
        files: list[Path] = []
        for pat in args.files:
            files.extend(_glob_files(pat))
        return files, f"--files ({len(args.files)} patterns)"
    if args.staged:
        return _git_staged_files(), "--staged"
    if args.all:
        # Default --all selection: all files we know how to check.
        files = []
        files.extend((TECHNICAL_ROOT / "extenbooks").glob("*.xlsx"))
        files.extend((TECHNICAL_ROOT / "chopped").glob("*.csv"))
        files.extend((TECHNICAL_ROOT / "research").glob("*_research.json"))
        files.extend((TECHNICAL_ROOT / "code").rglob("*.py"))
        if REGISTRY_PATH.exists():
            files.append(REGISTRY_PATH)
        return files, "--all"
    # No flag: auto-select.
    if _has_git_repo():
        return _git_staged_files(), "auto:staged"
    return discover_files(argparse.Namespace(files=None, staged=False, all=True))


# ---------------------------------------------------------------------------
# Registry helpers (cached load)
# ---------------------------------------------------------------------------
_registry_cache: dict | None = None

def load_registry() -> dict | None:
    global _registry_cache
    if _registry_cache is not None:
        return _registry_cache
    if not REGISTRY_PATH.exists():
        return None
    try:
        with REGISTRY_PATH.open("r", encoding="utf-8") as f:
            _registry_cache = json.load(f)
        return _registry_cache
    except (OSError, json.JSONDecodeError) as e:
        print(f"WARN: failed to load registry: {e}", file=sys.stderr)
        return None


def _rel(p: Path) -> str:
    try:
        return str(p.relative_to(PROJECT_ROOT)).replace("\\", "/")
    except ValueError:
        return str(p)


# ---------------------------------------------------------------------------
# Decision 0001 — extenbook 4-sheet canonical layout
# ---------------------------------------------------------------------------
def check_decision_0001(files: Iterable[Path], rpt: Report) -> None:
    xlsx_files = [p for p in files if p.suffix.lower() == ".xlsx"
                  and "extenbooks" in p.parts]
    if not xlsx_files:
        return
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        rpt.add("Decision-0001", "WARN", "<openpyxl>",
                "openpyxl not installed; cannot verify extenbook sheets")
        return
    from openpyxl import load_workbook
    for p in xlsx_files:
        try:
            wb = load_workbook(str(p), read_only=True, data_only=True)
            sheets = list(wb.sheetnames)
            wb.close()
        except Exception as e:
            rpt.add("Decision-0001", "FAIL", _rel(p),
                    f"failed to open extenbook: {e}")
            continue
        if sheets == CANONICAL_EXTENBOOK_SHEETS or sheets == ALLOWED_EXTENBOOK_SHEETS_5:
            rpt.add("Decision-0001", "PASS", _rel(p),
                    f"canonical sheets ({len(sheets)})")
        else:
            rpt.add("Decision-0001", "FAIL", _rel(p),
                    f"non-canonical sheets: {sheets} "
                    f"(expected {CANONICAL_EXTENBOOK_SHEETS} "
                    f"or {ALLOWED_EXTENBOOK_SHEETS_5})")


# ---------------------------------------------------------------------------
# Decision 0002 — validation.reference_values required
# ---------------------------------------------------------------------------
def check_decision_0002(files: Iterable[Path], rpt: Report) -> None:
    if not any(p.name == "series_registry.json" for p in files):
        return
    reg = load_registry()
    if reg is None:
        rpt.add("Decision-0002", "FAIL", _rel(REGISTRY_PATH),
                "registry not loadable")
        return
    series = reg.get("series", {})
    if not series:
        rpt.add("Decision-0002", "WARN", _rel(REGISTRY_PATH),
                "no series in registry")
        return
    missing = []
    for sid, block in series.items():
        validation = block.get("validation") or {}
        ref = validation.get("reference_values")
        derived = validation.get("derived_statistics")
        # Allow empty reference_values only if derived_statistics is non-empty
        # (per Decision 0008's multi-column summary case).
        if not ref and not derived:
            missing.append(sid)
    if missing:
        rpt.add("Decision-0002", "FAIL", _rel(REGISTRY_PATH),
                f"{len(missing)} series missing validation.reference_values "
                f"and derived_statistics: {missing[:10]}"
                + ("..." if len(missing) > 10 else ""))
    else:
        rpt.add("Decision-0002", "PASS", _rel(REGISTRY_PATH),
                f"all {len(series)} series carry validation anchors")


# ---------------------------------------------------------------------------
# Decision 0003 — extension binary invariant
# ---------------------------------------------------------------------------
def check_decision_0003(files: Iterable[Path], rpt: Report) -> None:
    if not any(p.name == "series_registry.json" for p in files):
        return
    reg = load_registry()
    if reg is None:
        return
    series = reg.get("series", {})
    fail_ext_null_validated = []
    warn_ext_pop_no_keyword = []
    for sid, block in series.items():
        ext = block.get("extension")
        status = (block.get("status") or "").lower()
        if ext is None and "validated_book_and_extension" in status:
            fail_ext_null_validated.append(sid)
        elif ext is not None and "extension" not in status:
            warn_ext_pop_no_keyword.append(sid)
    if fail_ext_null_validated:
        rpt.add("Decision-0003", "FAIL", _rel(REGISTRY_PATH),
                f"{len(fail_ext_null_validated)} series claim "
                f"validated_book_and_extension with extension=null: "
                f"{fail_ext_null_validated[:10]}")
    if warn_ext_pop_no_keyword:
        rpt.add("Decision-0003", "WARN", _rel(REGISTRY_PATH),
                f"{len(warn_ext_pop_no_keyword)} series have extension block "
                f"but status lacks 'extension' keyword "
                f"(legitimate if extension is planned but not yet validated): "
                f"{warn_ext_pop_no_keyword[:10]}")
    if not fail_ext_null_validated and not warn_ext_pop_no_keyword:
        rpt.add("Decision-0003", "PASS", _rel(REGISTRY_PATH),
                f"all {len(series)} series satisfy extension binary invariant")


# ---------------------------------------------------------------------------
# Decision 0004 — L01/P02/V03 compact naming
# ---------------------------------------------------------------------------
_NAME_RE = re.compile(r"^(L01|P02|V03|M04|A05|O06|S00)_([A-Z]+\d+[A-Z]?)(?:_[A-Za-z0-9_]+)?\.py$")

def check_decision_0004(files: Iterable[Path], rpt: Report) -> None:
    py_files = [p for p in files if p.suffix == ".py"
                and p.name.startswith(PHASE_PREFIXES)
                and "code" in p.parts]
    if not py_files:
        return
    bad = []
    for p in py_files:
        if not _NAME_RE.match(p.name):
            bad.append(_rel(p))
    if bad:
        for b in bad:
            rpt.add("Decision-0004", "FAIL", b,
                    "filename does not match <phase>_<SID>[_<suffix>].py pattern")
    else:
        rpt.add("Decision-0004", "PASS",
                f"<{len(py_files)} phase scripts>",
                "all phase-prefixed scripts match compact naming pattern")


# ---------------------------------------------------------------------------
# Decision 0005 — chopped wide format
# ---------------------------------------------------------------------------
def check_decision_0005(files: Iterable[Path], rpt: Report) -> None:
    csv_files = [p for p in files if p.suffix.lower() == ".csv"
                 and "chopped" in p.parts]
    if not csv_files:
        return
    import csv as _csv
    for p in csv_files:
        try:
            with p.open("r", encoding="utf-8", newline="") as f:
                reader = _csv.reader(f)
                rows = []
                for i, row in enumerate(reader):
                    rows.append(row)
                    if i >= 2:
                        break
        except OSError as e:
            rpt.add("Decision-0005", "FAIL", _rel(p), f"unreadable: {e}")
            continue
        if len(rows) < 3:
            rpt.add("Decision-0005", "FAIL", _rel(p),
                    f"fewer than 3 rows present ({len(rows)})")
            continue
        row1, row2, row3 = rows[0], rows[1], rows[2]
        errors = []
        # Row 1 — metadata row. Project pattern (Decision 0005): the first cell
        # is empty (or a '#'-style marker) and a later cell carries a quoted
        # metadata string with semicolon-separated descriptors. Accept either
        # the legacy '#'-prefixed form *or* the RMWND empty-first-cell form.
        row1_first = (row1[0] if row1 else "").strip()
        row1_joined = ",".join(row1)
        looks_like_metadata = (
            row1_joined.startswith("#")
            or "#" in row1_joined
            or (row1_first == "" and len(row1) >= 2
                and (";" in row1[1] or "=" in row1[1] or "period" in row1[1].lower()))
        )
        if not looks_like_metadata:
            errors.append(
                f"row 1 not recognizable as metadata "
                f"(expected leading '#' or empty-first-cell + descriptor; got: "
                f"{row1_joined[:60]!r})"
            )
        # Row 2 — column-ID row. First cell must be 'Year' (or a year-like header).
        row2_first = (row2[0] if row2 else "").strip().lower()
        if not row2_first.startswith("year"):
            errors.append(
                f"row 2 first cell is not 'Year' (got: {row2_first!r})"
            )
        # Row 3 — first data row. First cell must be a 4-digit year.
        row3_first = (row3[0] if row3 else "").strip()
        if not re.match(r"^\d{4}$", row3_first):
            errors.append(
                f"row 3 first cell is not a 4-digit year (got: {row3_first!r})"
            )
        if errors:
            rpt.add("Decision-0005", "FAIL", _rel(p), "; ".join(errors))
        else:
            rpt.add("Decision-0005", "PASS", _rel(p), "wide format ok")


# ---------------------------------------------------------------------------
# Decision 0006 — code is source of truth (heuristic documentation check)
# ---------------------------------------------------------------------------
def check_decision_0006(files: Iterable[Path], rpt: Report) -> None:
    # Always-PASS: this is a meta-principle enforced by anu-doctor and human
    # review, not by a syntactic check. We surface it here so the hook reports
    # all eight decisions consistently.
    rpt.add("Decision-0006", "PASS", "<meta>",
            "code-as-source-of-truth is a meta-principle; "
            "enforced by anu-doctor + human review")


# ---------------------------------------------------------------------------
# Decision 0007 — canonical verbatim schema in research JSONs
# ---------------------------------------------------------------------------
_REQUIRED_VQ_FIELDS = ("entry_type", "source_ref")

def _is_canonical_vq(entry: dict) -> tuple[bool, str]:
    if entry.get("entry_type") != "verbatim_quote":
        return False, "entry_type != 'verbatim_quote'"
    if not entry.get("source_ref"):
        return False, "missing/empty source_ref"
    # Content text must live in either `content` or `verbatim_quote`.
    if not (entry.get("content") or entry.get("verbatim_quote")):
        return False, "neither content nor verbatim_quote populated"
    return True, ""


def _loader_visible_quotes(data: dict) -> int:
    """Count quotes the viz loader will surface (schema A: verbatim_quotes[]
    entries carrying a non-empty ``text`` — mirrors
    Technical/viz/data_loader._load_verbatim_quote)."""
    quotes = data.get("verbatim_quotes") or []
    if not isinstance(quotes, list):
        return 0
    return sum(1 for q in quotes if isinstance(q, dict) and q.get("text"))


def check_decision_0007(files: Iterable[Path], rpt: Report) -> None:
    json_files = [p for p in files
                  if p.suffix == ".json" and p.name.endswith("_research.json")
                  and "research" in p.parts]
    if not json_files:
        return
    reg = load_registry()
    reg_series = (reg or {}).get("series", {})
    for p in json_files:
        try:
            with p.open("r", encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, json.JSONDecodeError) as e:
            rpt.add("Decision-0007", "FAIL", _rel(p), f"unreadable: {e}")
            continue
        entries = data.get("entries") or []
        canonical = [e for e in entries
                     if isinstance(e, dict) and _is_canonical_vq(e)[0]]
        if len(canonical) >= 3:
            rpt.add("Decision-0007", "PASS", _rel(p),
                    f"{len(canonical)} canonical verbatim_quote entries")
        else:
            rpt.add("Decision-0007", "FAIL", _rel(p),
                    f"only {len(canonical)} canonical verbatim_quote entries "
                    f"(need >=3 per Decision 0007)")
        # HARD loader-visibility check (Tier-A W1a, 2026-07-08): every
        # publish:true series must ALSO carry schema-A verbatim_quotes[] with
        # a non-empty `text`, or the website author quote renders empty. The
        # schema-B-only tolerance above let 14 published series ship without
        # a visible quote (A4 audit, 2026-07-07).
        sid = p.name[:-len("_research.json")]
        meta = reg_series.get(sid)
        if meta is not None and meta.get("publish") is True:
            n_visible = _loader_visible_quotes(data)
            if n_visible == 0:
                rpt.add("Decision-0007", "FAIL", _rel(p),
                        "publish:true series has NO loader-visible quote "
                        "(schema-A verbatim_quotes[] with non-empty 'text'); "
                        "website author quote would render empty")
            else:
                rpt.add("Decision-0007", "PASS", _rel(p),
                        f"{n_visible} loader-visible schema-A quotes "
                        f"(published-series hard check)")


# ---------------------------------------------------------------------------
# Decision 0008 — reference_values year-keyed scalars only
# ---------------------------------------------------------------------------
def check_decision_0008(files: Iterable[Path], rpt: Report) -> None:
    if not any(p.name == "series_registry.json" for p in files):
        return
    reg = load_registry()
    if reg is None:
        return
    series = reg.get("series", {})
    offenders: list[str] = []
    for sid, block in series.items():
        validation = block.get("validation") or {}
        ref = validation.get("reference_values") or {}
        for k, v in ref.items():
            try:
                int(k)
            except (TypeError, ValueError):
                offenders.append(f"{sid}:{k!r}=non-year-key")
                continue
            if not isinstance(v, (int, float)) or (
                isinstance(v, float) and not math.isfinite(v)
            ):
                offenders.append(f"{sid}:{k}=non-finite-scalar({v!r})")
    if offenders:
        rpt.add("Decision-0008", "FAIL", _rel(REGISTRY_PATH),
                f"{len(offenders)} reference_values violations: "
                f"{offenders[:10]}" + ("..." if len(offenders) > 10 else ""))
    else:
        rpt.add("Decision-0008", "PASS", _rel(REGISTRY_PATH),
                f"all reference_values keys parse as int years, "
                f"values are finite scalars across {len(series)} series")


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
ALL_CHECKS = [
    ("Decision-0001", check_decision_0001),
    ("Decision-0002", check_decision_0002),
    ("Decision-0003", check_decision_0003),
    ("Decision-0004", check_decision_0004),
    ("Decision-0005", check_decision_0005),
    ("Decision-0006", check_decision_0006),
    ("Decision-0007", check_decision_0007),
    ("Decision-0008", check_decision_0008),
]


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        description="Pre-commit check for RMWND Decisions 0001-0008.",
    )
    g = ap.add_mutually_exclusive_group()
    g.add_argument("--all", action="store_true",
                   help="check the whole project (default if no git repo)")
    g.add_argument("--staged", action="store_true",
                   help="check only git-staged files (default if git repo)")
    ap.add_argument("--files", nargs="+", default=None,
                    help="glob patterns relative to project root")
    ap.add_argument("--verbose", "-v", action="store_true",
                    help="show PASS lines too (default: WARN/FAIL only)")
    args = ap.parse_args(argv)

    files, mode = discover_files(args)
    rpt = Report()

    # Decision 0002/0003/0008 always run when registry is in scope.
    # Force registry into the file list for whole-project / staged runs that
    # didn't touch it, so the always-on checks still report a PASS line.
    if (args.all or mode in ("auto:staged", "--staged")) and REGISTRY_PATH.exists():
        if REGISTRY_PATH not in files:
            files = list(files) + [REGISTRY_PATH]

    for _, fn in ALL_CHECKS:
        fn(files, rpt)

    counts = rpt.counts()
    print(f"\nRMWND precommit check ({mode}) — "
          f"{len(files)} file(s) considered")
    print("-" * 100)
    for r in rpt.results:
        if r.level == "PASS" and not args.verbose:
            continue
        print(r.fmt())
    print("-" * 100)
    print(f"PASS={counts.get('PASS',0)}  "
          f"WARN={counts.get('WARN',0)}  "
          f"FAIL={counts.get('FAIL',0)}")
    return rpt.exit_code()


if __name__ == "__main__":
    sys.exit(main())
